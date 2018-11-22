import random
from openpyxl import load_workbook
import xlrd
from prettytable import PrettyTable
from tkinter import *

grploc = r"C:\Users\sai sudhamsa\Desktop\TImeTable Project\Groups.xlsx"
tchrloc=r"C:\Users\sai sudhamsa\Desktop\TImeTable Project\Teachers.xlsx"
stuloc=r"C:\Users\sai sudhamsa\Desktop\TImeTable Project\Students.xlsx"
courseloc=r"C:\Users\sai sudhamsa\Desktop\TImeTable Project\Courses.xlsx"

alpha=['A','B','C','D','E','F','G','H','I','J','K','L','M']

grpwb = xlrd.open_workbook(grploc)
stuwb = xlrd.open_workbook(stuloc)
tchrwb = xlrd.open_workbook(tchrloc)
crseswb = xlrd.open_workbook(courseloc)

grp = grpwb.sheet_by_index(0)
stu = stuwb.sheet_by_index(0)
tchr = tchrwb.sheet_by_index(0)
crses = crseswb.sheet_by_index(0)

grpwb1 = load_workbook(grploc)
stuwb1 = load_workbook(stuloc)
tchrwb1 = load_workbook(tchrloc)
crseswb1 = load_workbook(courseloc)

grp1 = grpwb1['Sheet1']
stu1 = stuwb1['Sheet1']
tchr1 = tchrwb1['Sheet1']
crses1 =crseswb1['Sheet1']
#CLASSES

def createTimetable():
    print("Started")
    class Year:
        def __init__(self, name, courses):
            self.courses = courses
            self.name = name
            self.classes = [[None for i in range(6)] for j in range(5)]

        def fillrandom(self):
            for i in self.courses:
                l = intersection(self, i.teacher)
                # if(i==[]):
                l = shuffle(l)
                #print(len(l))
                count = 0
                while (i.hours < i.no_of_classes):
                    self.classes[l[count][0]][l[count][1]] = i
                    i.teacher.addclass()
                    i.addclass()
                    i.teacher.classes[l[count][0]][l[count][1]] = i
                    count += 1
                # print(self.classes)

        def assignBlock(self, c):
            a = random.randint(0, 4)
            b = random.randint(0, 4)
            if (self.classes[a][b] is None and self.classes[a + 1][b] is None):
                if (c.teacher.isFree((a, b)) and c.teacher.isFree((a + 1, b))):
                    self.classes[a][b] = c
                    self.classes[a + 1][b] = c
                    c.teacher.addclass(2)
                    c.teacher.classes[a][b] = c
                    c.teacher.classes[a + 1][b] = c
                    c.addclass(2)

        def fixCourse(self, c, time):
            self.classes[time[0]][time[1]] = c
            c.teacher.addclass()
            c.teacher.classes[time[0]][time[1]] = c
            c.addclass()

        def assignTeachers(self):
                i.assignTeacher()

    def intersection(a, b):
        ans = []
        m = 0
        for i in a.classes:
            n = 0
            for j in i:
                if (j is None and b.isFree((m, n))):
                    ans.append((m, n))
                n += 1
            m += 1
        return ans

    def shuffle(x):
        l = x.copy()
        ans = [x[0]]
        del l[0]
        while (True):
            count = 0
            while (count < len(l)):
                if (l[count][0] != ans[-1][0]):
                    ans.append(l[count])
                    del l[count]
                    count -= 1
                count += 1
            if (check(l)):
                ans.extend(l)
                break
        shuffle1(ans)
        return ans

    def shuffle1(x):
        c = []
        i = 0
        while (i < len(x)):
            j = i + 1
            c.clear()
            while (j < len(x)):
                if (x[i][0] == x[j][0]):
                    c.append(j)
                j += 1
            if (c != []):
                k = random.choice(c)
                temp = x[i]
                x[i] = x[k]
                x[k] = temp
            i += 1

    def check(l):
        for i in l:
            if (i[0] != l[0][0]):
                return False
        return True

    class Teacher:
        def __init__(self, name, courses):
            self.name = name
            self.courses = courses
            self.hours = 0
            self.freeslots = list()
            self.classes = [[None for i in range(6)] for j in range(5)]

        def addclass(self, n=1):
            self.hours += n

        def isFree(self, time):
            if (self.classes[time[0]][time[1]] == None and time not in self.freeslots):
                return True
            return False

        def addFreeslot(self, time):
            self.freeslots.append(time)

    class course:
        def __init__(self, name, no_of_classes, year):
            self.name = name
            self.year = year
            self.no_of_classes = no_of_classes
            self.teacher = None
            self.hours = 0

        def addclass(self, n=1):
            self.hours += n

        def assignTeacher(self):
            count = 0
            teacherlist = []

            for i in tchrlist:
                if (self in i.courses):
                    count += 1
                    teacherlist.append(i)
            if (count == 1):
                self.teacher = teacherlist[0]
            elif (count > 1):
                # print(teacherlist[0].name,teacherlist[1].name)
                minimum = teacherlist[0].hours
                self.teacher = teacherlist[0]
                for i in teacherlist:
                    if (i.hours < minimum):
                        self.teacher = i
            else:
                print("Error input" + self.name)

    def printtable(periods, bool):
        table = PrettyTable(['DAY', '8.30-9.25', '9.30-10.25', '10.40-11.35', '11.40-12.35', '12.40-1.30', '1.30-2.25'])
        c = []
        D = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        count = 0
        for i in periods:
            c.clear()
            c.append(D[count])
            for j in i:
                if (j is not None):
                    if (bool):
                        c.append(j.name + "\n" + j.year)
                    else:
                        c.append(j.name + "\n" + j.teacher.name)
                else:
                    c.append("  ")
            table.add_row(c)
            count += 1
        return table

    courselist = []
    count = 1
    while(count<crses.nrows):
        courselist.append(course(crses.cell_value(count, 0), int(crses.cell_value(count, 1)), crses.cell_value(count, 2)))
        count+=1

    def findCourse(name):
        for i in courselist:
            if (i.name == name):
                return i

    grplist = []
    count = 1
    while(count<grp.nrows):
        subjs = []
        count1 = 2
        while (True):
            if (count1 < grp.ncols and grp.cell_value(count, count1) != ""):
                subjs.append(findCourse(grp.cell_value(count, count1)))
            else:
                break
            count1 += 1
        grplist.append(Year(grp.cell_value(count, 0), subjs))
        count += 1

    tchrlist = []
    count = 1
    while(count<tchr.nrows):
        subjs = []
        count1 = 3
        while (True):
            if (count1 < tchr.ncols and tchr.cell_value(count, count1) != ""):
                subjs.append(findCourse(tchr.cell_value(count, count1)))
            else:
                break
            count1 += 1
        tchrlist.append(Teacher(tchr.cell_value(count, 0), subjs))
        count+=1

    for i in grplist:
        i.assignTeachers()
        print("teachers assigned")
        i.fillrandom()
        print("time table filled")

    grpwb2 = load_workbook(grploc)
    tchrwb2 = load_workbook(tchrloc)
    grp2 = grpwb2['Sheet1']
    tchr2 = tchrwb2['Sheet1']
    count=2
    for i in tchrlist:
        tchr2['C'+str(count)]="\'"+str(printtable(i.classes,True))
        count+=1
    tchrwb2.save(tchrloc)
    count=2
    for i in grplist:
        grp2['B'+str(count)]="\'"+str(printtable(i.classes, False))
        count+=1
    grpwb2.save(grploc)
    print("Ended")
#___________________________________________________________________________________________
#Database

# To open Workbook
#-------------------------------------------------------------------------------------------------------------
#GUI
def check(a,b):
    if(a=="ok" and b=="ok"):
        return True
    else:
        pass

def main():
    student_frame.grid_forget()
    stuSign_frame.grid_forget()
    stuPage_frame.grid_forget()
    teacher_frame.grid_forget()
    teaSign_frame.grid_forget()
    teaPage_frame.grid_forget()
    admin_frame.grid_forget()
    adminPage_frame.grid_forget()
    main_frame.grid_forget()
    #year_frame.grid_forget()
    cource_frame.grid_forget()
    #forSubjects_Frame.grid_forget()

    main_frame.grid(column=0, row=0, padx=20, pady=5)

def in_main():
    button1 = Button(main_frame, text="Student", fg="Blue", command=student)
    button1.place(x=400, y=100)
    button1.config(width=10, height=5)

    button2 = Button(main_frame, text="Teacher", fg="Green", command=teacher)
    button2.place(x=400, y=300)
    button2.config(width=10, height=5)

def student():
    stuSign_frame.grid_forget()
    stuPage_frame.grid_forget()
    teacher_frame.grid_forget()
    teaSign_frame.grid_forget()
    teaPage_frame.grid_forget()
    admin_frame.grid_forget()
    adminPage_frame.grid_forget()
    main_frame.grid_forget()
    #year_frame.grid_forget()
    cource_frame.grid_forget()
    #forSubjects_Frame.grid_forget()

    student_frame.grid(column=0, row=0, padx=20, pady=5)

def findYear(count):
    if(stu.cell_value(count,2)=="Year1"):
        return 2
    elif(stu.cell_value(count,2)=="Year2"):
        return 3
    elif (stu.cell_value(count,2) == "Year3"):
        return 4
    elif (stu.cell_value(count,2) == "Year4"):
        return 5

def in_student():
    def stuPage():
        count = 1
        while (count < stu.nrows):
            if (username.get() == stu.cell_value(count, 0)):
                if (password.get() == stu.cell_value(count, 1)):
                    student_frame.grid_forget()
                    stuSign_frame.grid_forget()
                    teacher_frame.grid_forget()
                    teaSign_frame.grid_forget()
                    teaPage_frame.grid_forget()
                    admin_frame.grid_forget()
                    adminPage_frame.grid_forget()
                    main_frame.grid_forget()
                    cource_frame.grid_forget()
                    stuPage_frame.grid(column=0, row=0, padx=20, pady=5)
                    label1 = Label(stuPage_frame, text=grp1['B'+str(findYear(count))], font = "Courier 12")
                    label1.pack()
                    main = Button(stuPage_frame, text="Back", command=student)
                    main.pack()
            count += 1

    a = Label(student_frame, text="username", font="Courier 20").grid(row=0, column=0)
    b = Label(student_frame, text="password", font="Courier 20").grid(row=1, column=0)
    username = Entry(student_frame)
    username.grid(row=0, column=1)
    password = Entry(student_frame, show="*")
    password.grid(row=1, column=1)
    c = Button(student_frame, text="Login",command=stuPage)
    c.grid(row=2, column=1)

    d = Button(student_frame, text="Back", command = main)
    d.grid(row=2 , column=0)

def teacher():
    student_frame.grid_forget()
    stuSign_frame.grid_forget()
    stuPage_frame.grid_forget()
    teaSign_frame.grid_forget()
    teaPage_frame.grid_forget()
    admin_frame.grid_forget()
    adminPage_frame.grid_forget()
    main_frame.grid_forget()
    #year_frame.grid_forget()
    cource_frame.grid_forget()
    #forSubjects_Frame.grid_forget()

    teacher_frame.grid(column=0, row=0, padx=20, pady=5)

def in_teacher():
    def teaPage():
        count = 1
        while (count < tchr.nrows):
            if (username.get() == tchr.cell_value(count, 0)):
                if (password.get() == tchr.cell_value(count, 1)):
                    student_frame.grid_forget()
                    stuSign_frame.grid_forget()
                    stuPage_frame.grid_forget()
                    teacher_frame.grid_forget()
                    teaSign_frame.grid_forget()
                    admin_frame.grid_forget()
                    adminPage_frame.grid_forget()
                    main_frame.grid_forget()
                    cource_frame.grid_forget()
                    teaPage_frame.grid(column=0, row=0, padx=20, pady=5)
                    label1 = Label(teaPage_frame, text=tchr.cell_value(count,2), font = "Courier 12")
                    label1.pack()
                    main = Button(teaPage_frame, text="Back", command=teacher)
                    main.pack()
            count += 1

    a = Label(teacher_frame, text="username", font="Courier 20").grid(row=0, column=0)
    b = Label(teacher_frame, text="password", font="Courier 20").grid(row=1, column=0)
    username = Entry(teacher_frame)
    username.grid(row=0, column=1)
    password = Entry(teacher_frame, show="*")
    password.grid(row=1, column=1)
    c = Button(teacher_frame, text="Login", command=teaPage)
    c.grid(row=2, column=1)

    d = Button(teacher_frame, text="Back", command=main)
    d.grid(row=2, column=0)


def in_cource():
    def findfree(row):
        count = 2
        while(count < grp1.max_column and grp1[alpha[count] + str(row)].value != ""):
            count+=1
        return count

    def add():
        def addCourse():
            name = cource_entry.get()
            no_of_classes = classes_entry.get()
            year = year_entry.get()
            col = crses1.max_row + 1
            crses1['A' + str(col)] = name
            crses1['B' + str(col)] = no_of_classes
            crses1['C' + str(col)] = year
            if (year == "Year1"):
                grp1[alpha[findfree(2)] + str(2)] = name
            elif (year == "Year2"):
                grp1[alpha[findfree(3)] + str(3)] = name
            elif (year == "Year3"):
                grp1[alpha[findfree[4]] + str(4)] = name
            elif (year == "Year3"):
                grp1[alpha[findfree[5]] + str(5)] = name
            crseswb1.save(courseloc)
            grpwb1.save(grploc)
        cource = Label(cource_frame, text="Cource").pack()
        cource_entry = Entry(cource_frame)
        cource_entry.pack()

        year = Label(cource_frame, text="Year").pack()
        year_entry = Entry(cource_frame)
        year_entry.pack()

        classes = Label(cource_frame, text="No.of classes").pack()
        classes_entry = Entry(cource_frame)
        classes_entry.pack()

        done = Button(cource_frame, text="Done", command=addCourse).pack()

    c = Button(cource_frame, text="Add", command=add)
    c.pack()

    main = Button(cource_frame, text="Back", command=in_adminCource)
    main.pack()

def in_stuSign():

    a = Label(stuSign_frame, text="Full Name", font="Courier 20").grid(row=0, column=0)
    c = Label(stuSign_frame, text="Password", font="Courier 20").grid(row=1, column=0)
    b = Label(stuSign_frame, text="Year", font="Courier 20").grid(row=2, column=0)

    def addStudent():
        name=full_name.get()
        psswd=password.get()
        year=year_entry.get()
        col=stu1.max_row + 1
        stu1['A' + str(col)] = name
        stu1['B' + str(col)] = psswd
        stu1['C' + str(col)] = year
        stuwb1.save(stuloc)

    full_name = Entry(stuSign_frame)
    full_name.grid(row=0, column=1)
    password = Entry(stuSign_frame, show='*')
    password.grid(row=1, column=1)
    year_entry = Entry(stuSign_frame)
    year_entry.grid(row=2, column=1)


    submit_button = Button(stuSign_frame, text="Submit",command=addStudent)
    submit_button.grid(row=3, columnspan=2)

    main= Button(stuSign_frame, text="Back",command=in_adminCource)
    main.grid(row=4,column=1)

def in_teaSign():
    all_subjects = []

    def add_subject():
        def add():
            all_subjects.append(entry.get())
        subject = Label(teaSign_frame, text="Subject Name :").pack()
        entry = Entry(teaSign_frame)
        entry.pack()
        done = Button(teaSign_frame, text ="Done", command=add).pack()

    def addTeacher():
        name=full_name.get()
        password=password_Entry.get()
        col=tchr1.max_row + 1
        tchr1['A' + str(col)] = name
        tchr1['B' + str(col)] = password
        count = 0
        while (count < len(all_subjects)):
            tchr1[alpha[count + 3] + str(col)] = all_subjects[count]
            count += 1
        tchrwb1.save(tchrloc)

    button = Button(teaSign_frame, text="Add subject", command=add_subject)
    button.pack()

    a = Label(teaSign_frame, text="Full Name", font="Courier 20").pack()

    full_name = Entry(teaSign_frame)
    full_name.pack()

    b = Label(teaSign_frame, text="Password", font="Courier 20").pack()

    password_Entry = Entry(teaSign_frame, show='*')
    password_Entry.pack()

    submit = Button(teaSign_frame, text="Submit", command=addTeacher).pack()

    main = Button(teaSign_frame, text="Back", command=in_adminTs).pack()


def in_adminCource():
    def cource():
        if (check(username.get(), password.get())):
            student_frame.grid_forget()
            stuSign_frame.grid_forget()
            stuPage_frame.grid_forget()
            teacher_frame.grid_forget()
            teaSign_frame.grid_forget()
            teaPage_frame.grid_forget()
            admin_frame.grid_forget()
            adminPage_frame.grid_forget()
            main_frame.grid_forget()
            # year_frame.grid_forget()
            # forSubjects_Frame.grid_forget()

            cource_frame.grid(column=0, row=0, padx=20, pady=5)
        else:
            student_frame.grid_forget()
            stuSign_frame.grid_forget()
            stuPage_frame.grid_forget()
            teacher_frame.grid_forget()
            teaPage_frame.grid_forget()
            teaSign_frame.grid_forget()
            adminPage_frame.grid_forget()
            main_frame.grid_forget()
            # year_frame.grid_forget()
            cource_frame.grid_forget()
            # forSubjects_Frame.grid_forget()

            admin_frame.grid(column=0, row=0, padx=20, pady=5)



    student_frame.grid_forget()
    stuSign_frame.grid_forget()
    stuPage_frame.grid_forget()
    teacher_frame.grid_forget()
    teaPage_frame.grid_forget()
    teaSign_frame.grid_forget()
    adminPage_frame.grid_forget()
    main_frame.grid_forget()
    #year_frame.grid_forget()
    cource_frame.grid_forget()
    #forSubjects_Frame.grid_forget()

    admin_frame.grid(column=0, row=0, padx=20, pady=5)

    a = Label(admin_frame, text="username", font="Courier 20").grid(row=0, column=0)
    b = Label(admin_frame, text="password", font="Courier 20").grid(row=1, column=0)
    username = Entry(admin_frame)
    username.grid(row=0, column=1)
    password = Entry(admin_frame, show="*")
    password.grid(row=1, column=1)
    c = Button(admin_frame, text="Login", command=cource)
    c.grid(row=2, column=1)

    d = Button(admin_frame, text="Back", command=main)
    d.grid(row=2, column=0)

def in_adminTs():
    def tea_signup():
        if (check(username.get(), password.get())):
            student_frame.grid_forget()
            stuSign_frame.grid_forget()
            stuPage_frame.grid_forget()
            teacher_frame.grid_forget()
            teaPage_frame.grid_forget()
            admin_frame.grid_forget()
            adminPage_frame.grid_forget()
            main_frame.grid_forget()
            cource_frame.grid_forget()
            """ year_frame.grid_forget()
            forSubjects_Frame.grid_forget()"""

            teaSign_frame.grid(column=0, row=0, padx=20, pady=5)
        else:
            student_frame.grid_forget()
            stuSign_frame.grid_forget()
            stuPage_frame.grid_forget()
            teacher_frame.grid_forget()
            teaPage_frame.grid_forget()
            teaSign_frame.grid_forget()
            adminPage_frame.grid_forget()
            main_frame.grid_forget()
            # year_frame.grid_forget()
            cource_frame.grid_forget()
            # forSubjects_Frame.grid_forget()

            admin_frame.grid(column=0, row=0, padx=20, pady=5)

    student_frame.grid_forget()
    stuSign_frame.grid_forget()
    stuPage_frame.grid_forget()
    teacher_frame.grid_forget()
    teaPage_frame.grid_forget()
    teaSign_frame.grid_forget()
    adminPage_frame.grid_forget()
    main_frame.grid_forget()
    #year_frame.grid_forget()
    cource_frame.grid_forget()
    #forSubjects_Frame.grid_forget()

    admin_frame.grid(column=0, row=0, padx=20, pady=5)

    a = Label(admin_frame, text="username", font="Courier 20").grid(row=0, column=0)
    b = Label(admin_frame, text="password", font="Courier 20").grid(row=1, column=0)
    username = Entry(admin_frame)
    username.grid(row=0, column=1)
    password = Entry(admin_frame, show="*")
    password.grid(row=1, column=1)
    c = Button(admin_frame, text="Login", command=tea_signup)
    c.grid(row=2, column=1)

    d = Button(admin_frame, text="Back", command=main)
    d.grid(row=2, column=0)

def in_adminSs():
    def stu_signup():
        if (check(username.get(), password.get())):
            student_frame.grid_forget()
            stuPage_frame.grid_forget()
            teacher_frame.grid_forget()
            teaSign_frame.grid_forget()
            teaPage_frame.grid_forget()
            admin_frame.grid_forget()
            adminPage_frame.grid_forget()
            main_frame.grid_forget()
            # year_frame.grid_forget()
            cource_frame.grid_forget()
            # forSubjects_Frame.grid_forget()

            stuSign_frame.grid(column=0, row=0, padx=20, pady=5)
        else:
            student_frame.grid_forget()
            stuSign_frame.grid_forget()
            stuPage_frame.grid_forget()
            teacher_frame.grid_forget()
            teaPage_frame.grid_forget()
            teaSign_frame.grid_forget()
            adminPage_frame.grid_forget()
            main_frame.grid_forget()
            # year_frame.grid_forget()
            cource_frame.grid_forget()
            # forSubjects_Frame.grid_forget()

            admin_frame.grid(column=0, row=0, padx=20, pady=5)

    student_frame.grid_forget()
    stuSign_frame.grid_forget()
    stuPage_frame.grid_forget()
    teacher_frame.grid_forget()
    teaPage_frame.grid_forget()
    teaSign_frame.grid_forget()
    adminPage_frame.grid_forget()
    main_frame.grid_forget()
    #year_frame.grid_forget()
    cource_frame.grid_forget()
    #forSubjects_Frame.grid_forget()

    admin_frame.grid(column=0, row=0, padx=20, pady=5)

    a = Label(admin_frame, text="username", font="Courier 20").grid(row=0, column=0)
    b = Label(admin_frame, text="password", font="Courier 20").grid(row=1, column=0)
    username = Entry(admin_frame)
    username.grid(row=0, column=1)
    password = Entry(admin_frame, show="*")
    password.grid(row=1, column=1)
    c = Button(admin_frame, text="Login", command=stu_signup)
    c.grid(row=2, column=1)

    d = Button(admin_frame, text="Back", command=main)
    d.grid(row=2, column=0)

window_width = 750
window_heigth = 725

application = Tk()
application.geometry('800x750')
application.title("Time Table")

main_menu = Menu(application)
application.config(menu=main_menu)
sub_menu = Menu(main_menu)

main_menu.add_cascade(label="Menu", menu=sub_menu)

"""sub_menu.add_command(label="Year", command=in_adminYear)
sub_menu.add_separator()"""
sub_menu.add_command(label="Cource", command=in_adminCource)
sub_menu.add_separator()
sub_menu.add_command(label="Create Timetable", command=createTimetable)
sub_menu.add_separator()
sub_menu.add_command(label="Student", command=in_adminSs)
sub_menu.add_command(label="Teacher", command=in_adminTs)

main_frame = Frame(application, width=window_width, height=window_heigth)
main_frame['borderwidth'] = 20
main_frame['relief'] = 'sunken'
main_frame.grid(column=0, row=0, padx=20, pady=5)

student_frame = Frame(application, width=window_width, height=window_heigth)
student_frame['borderwidth'] = 20
student_frame['relief'] = 'sunken'
student_frame.grid(column=0, row=0, padx=20, pady=5)

teacher_frame = Frame(application, width=window_width, height=window_heigth)
teacher_frame['borderwidth'] = 20
teacher_frame['relief'] = 'sunken'
teacher_frame.grid(column=0, row=0, padx=20, pady=5)

admin_frame = Frame(application, width=window_width, height=window_heigth)
admin_frame['borderwidth'] = 20
admin_frame['relief'] = 'sunken'
admin_frame.grid(column=0, row=0, padx=20, pady=5)

stuPage_frame = Frame(application, width=window_width, height=window_heigth)
stuPage_frame['borderwidth'] = 20
stuPage_frame['relief'] = 'sunken'
stuPage_frame.grid(column=0, row=0, padx=20, pady=5)

teaPage_frame = Frame(application, width=window_width, height=window_heigth)
teaPage_frame['borderwidth'] = 20
teaPage_frame['relief'] = 'sunken'
teaPage_frame.grid(column=0, row=0, padx=20, pady=5)

adminPage_frame = Frame(application, width=window_width, height=window_heigth)
adminPage_frame['borderwidth'] = 20
adminPage_frame['relief'] = 'sunken'
adminPage_frame.grid(column=0, row=0, padx=20, pady=5)

cource_frame = Frame(application, width=window_width, height=window_heigth)
cource_frame['borderwidth'] = 20
cource_frame['relief'] = 'sunken'
cource_frame.grid(column=0, row=0, padx=20, pady=5)

teaSign_frame = Frame(application, width=window_width, height=window_heigth)
teaSign_frame['borderwidth'] = 20
teaSign_frame['relief'] = 'sunken'
teaSign_frame.grid(column=0, row=0, padx=20, pady=5)

stuSign_frame = Frame(application, width=window_width, height=window_heigth)
stuSign_frame['borderwidth'] = 20
stuSign_frame['relief'] = 'sunken'
stuSign_frame.grid(column=0, row=0, padx=20, pady=5)

in_main()
in_student()
in_teacher()
#in_teaPage()
#in_year()
in_cource()
in_stuSign()
in_teaSign()

student_frame.grid_forget()
teacher_frame.grid_forget()
admin_frame.grid_forget()
stuPage_frame.grid_forget()
teaPage_frame.grid_forget()
adminPage_frame.grid_forget()
#year_frame.grid_forget()
cource_frame.grid_forget()
stuSign_frame.grid_forget()
teaSign_frame.grid_forget()

application.mainloop()